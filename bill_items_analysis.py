import pandas as pd
import openpyxl
from openpyxl import load_workbook
import json

def analyze_bill_items(file_path):
    """Analyze the actual bill items in the statutory format"""
    
    # Load the workbook
    wb = load_workbook(file_path)
    ws = wb['BILL QUANTITY']
    
    print("=== BILL ITEMS ANALYSIS ===")
    
    # Look for the actual bill items (seems to start around row 25-30 based on typical formats)
    print("\nSearching for bill item headers...")
    
    # Check rows for the actual bill table headers
    bill_headers_row = None
    for row in range(20, 35):  # Check rows 20-34
        cell_a = ws.cell(row=row, column=1).value
        cell_b = ws.cell(row=row, column=2).value
        if cell_a and "Unit" in str(cell_a) and cell_b and "Qty" in str(cell_b):
            bill_headers_row = row
            break
    
    if bill_headers_row:
        print(f"Bill table headers found at row {bill_headers_row}")
        
        # Print the header row
        print("Bill table headers:")
        header_row = []
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=bill_headers_row, column=col)
            header_row.append(cell.value)
            col_letter = openpyxl.utils.get_column_letter(col)
            print(f"{col_letter}{bill_headers_row}: {cell.value}")
        
        # Print first few actual bill items
        print(f"\nFirst 5 bill items (rows {bill_headers_row+1}-{bill_headers_row+5}):")
        for row in range(bill_headers_row+1, min(bill_headers_row+6, ws.max_row + 1)):
            item_data = []
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                item_data.append(cell.value)
            print(f"Row {row}: {item_data}")
            
        # Check formatting of bill items
        print(f"\nFormatting of bill items (row {bill_headers_row+1}):")
        sample_item_row = bill_headers_row + 1
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=sample_item_row, column=col)
            col_letter = openpyxl.utils.get_column_letter(col)
            alignment = "right" if col in [3, 6, 7] else "left"  # Numbers typically right-aligned
            print(f"{col_letter}{sample_item_row}: Font={cell.font.name}, Size={cell.font.size}, "
                  f"Bold={cell.font.bold}, Alignment={alignment}")
    else:
        print("Could not find bill table headers. Checking rows 25-40 for item data:")
        for row in range(25, 41):
            row_data = []
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                row_data.append(cell.value)
            print(f"Row {row}: {row_data}")

if __name__ == "__main__":
    try:
        analyze_bill_items("attached_assets/BILL AND DEVIATION APPROVED 02 APRIL 2005.xlsm")
    except Exception as e:
        print(f"Error analyzing file: {e}")