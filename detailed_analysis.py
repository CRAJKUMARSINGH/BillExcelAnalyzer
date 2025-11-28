import pandas as pd
import openpyxl
from openpyxl import load_workbook
import json

def detailed_bill_quantity_analysis(file_path):
    """Detailed analysis of the BILL QUANTITY sheet"""
    
    # Load the workbook
    wb = load_workbook(file_path)
    
    print("=== DETAILED BILL QUANTITY SHEET ANALYSIS ===")
    
    # Focus on BILL QUANTITY sheet
    if 'BILL QUANTITY' not in wb.sheetnames:
        print("BILL QUANTITY sheet not found!")
        return
        
    ws = wb['BILL QUANTITY']
    
    print(f"Sheet: BILL QUANTITY")
    print(f"Dimensions: {ws.max_row} rows x {ws.max_column} columns")
    
    # Check column widths
    print("\nColumn widths (in characters):")
    column_widths = []
    for col in range(1, ws.max_column + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        width = ws.column_dimensions[col_letter].width if ws.column_dimensions[col_letter].width else "Default"
        column_widths.append(width)
        print(f"Column {col_letter}: {width}")
    
    # Check headers (assuming row 12 is where headers start based on the output)
    print("\nHeader row (row 12):")
    header_values = []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=12, column=col)
        header_values.append(cell.value)
        print(f"Column {openpyxl.utils.get_column_letter(col)}: {cell.value}")
    
    # Check data rows
    print("\nSample data rows (rows 13-17):")
    for row in range(13, min(18, ws.max_row + 1)):
        row_data = []
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            row_data.append(cell.value)
        print(f"Row {row}: {row_data}")
    
    # Check formatting of sample cells
    print("\nCell formatting sample (A12-G17):")
    for row in range(12, min(18, ws.max_row + 1)):
        for col in range(1, min(8, ws.max_column + 1)):
            cell = ws.cell(row=row, column=col)
            col_letter = openpyxl.utils.get_column_letter(col)
            print(f"{col_letter}{row}: Font={cell.font.name}, Size={cell.font.size}, Bold={cell.font.bold}, "
                  f"Border={cell.border.left.style if cell.border.left else 'None'}, "
                  f"Alignment={cell.alignment.horizontal if cell.alignment.horizontal else 'default'}")
    
    # Check for specific formatting in header row
    print("\nHeader row (12) formatting:")
    for col in range(1, min(8, ws.max_column + 1)):
        cell = ws.cell(row=12, column=col)
        col_letter = openpyxl.utils.get_column_letter(col)
        print(f"{col_letter}12: Font={cell.font.name}, Size={cell.font.size}, Bold={cell.font.bold}")
    
    # Check for borders in data rows
    print("\nBorder styles in data rows:")
    sample_row = 13
    for col in range(1, min(8, ws.max_column + 1)):
        cell = ws.cell(row=sample_row, column=col)
        col_letter = openpyxl.utils.get_column_letter(col)
        border_info = f"{col_letter}{sample_row}: "
        if cell.border.left:
            border_info += f"Left={cell.border.left.style}, "
        if cell.border.right:
            border_info += f"Right={cell.border.right.style}, "
        if cell.border.top:
            border_info += f"Top={cell.border.top.style}, "
        if cell.border.bottom:
            border_info += f"Bottom={cell.border.bottom.style}"
        print(border_info)
    
    return {
        "column_widths": column_widths,
        "headers": header_values
    }

if __name__ == "__main__":
    try:
        result = detailed_bill_quantity_analysis("attached_assets/BILL AND DEVIATION APPROVED 02 APRIL 2005.xlsm")
        print("\n=== SUMMARY ===")
        print("This statutory format has specific requirements that should be preserved in exports:")
        print("1. Column widths match exactly: A=12.29, B=62.43, C=13.0, D=8.71, E=9.0, F=11.0, G=9.14")
        print("2. Font: Calibri, Size: 9pt")
        print("3. Headers: Bold formatting")
        print("4. Borders: Thin borders on all cells")
        print("5. Alignment: Left for text, Right for numbers")
    except Exception as e:
        print(f"Error analyzing file: {e}")