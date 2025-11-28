import pandas as pd
import openpyxl
from openpyxl import load_workbook
import json

def analyze_excel_format(file_path):
    """Analyze the structure and formatting of the statutory Excel file"""
    
    # Load the workbook
    wb = load_workbook(file_path)
    
    print("=== STATUTORY FORMAT ANALYSIS ===")
    print(f"File: {file_path}")
    print(f"Worksheets: {wb.sheetnames}")
    
    # Analyze each worksheet
    for sheet_name in wb.sheetnames:
        print(f"\n--- Worksheet: {sheet_name} ---")
        ws = wb[sheet_name]
        
        # Get basic info
        print(f"Dimensions: {ws.max_row} rows x {ws.max_column} columns")
        
        # Check first few rows for structure
        print("\nFirst 10 rows:")
        for row in range(1, min(11, ws.max_row + 1)):
            row_data = []
            for col in range(1, min(11, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)
                row_data.append(f"{cell.value}")
            print(f"Row {row}: {row_data}")
        
        # Check column widths
        print("\nColumn widths:")
        for col in range(1, min(11, ws.max_column + 1)):
            col_letter = openpyxl.utils.get_column_letter(col)
            width = ws.column_dimensions[col_letter].width if ws.column_dimensions[col_letter].width else "Default"
            print(f"Column {col_letter}: {width}")
            
        # Check some cell formatting
        print("\nSample cell formatting (A1-A5):")
        for row in range(1, min(6, ws.max_row + 1)):
            cell = ws.cell(row=row, column=1)
            print(f"A{row}: Font={cell.font.name}, Size={cell.font.size}, Bold={cell.font.bold}")
        
    # If there's a 'Bill Quantity' sheet, analyze it specifically
    if 'Bill Quantity' in wb.sheetnames:
        print("\n--- BILL QUANTITY SHEET ANALYSIS ---")
        ws = wb['Bill Quantity']
        
        # Check headers
        header_row = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
        print(f"Header row: {header_row}")
        
        # Check data rows
        print("\nFirst 5 data rows:")
        for row in range(2, min(7, ws.max_row + 1)):
            row_data = []
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                row_data.append(cell.value)
            print(f"Row {row}: {row_data}")

if __name__ == "__main__":
    try:
        analyze_excel_format("attached_assets/BILL AND DEVIATION APPROVED 02 APRIL 2005.xlsm")
    except Exception as e:
        print(f"Error analyzing file: {e}")